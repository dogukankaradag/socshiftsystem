"""On-call roster (Nöbetçi Listesi) — CRUD + XLSX/PDF ingest.

Dizayn:
  * `OnCallRoster` tablosunda ekibin (L2/MSSP) tarih aralıklı
    nöbet satırları tutulur.
  * Kullanıcılar XLSX veya PDF dosyası yüklerler; backend dosyayı parse edip
    normalize edilmiş satırlara çevirir ve bir `upload_batch` UUID'si ile kaydeder.
  * Orijinal dosya asla saklanmaz (kullanıcı isteği).

Parsing heuristics:
  * XLSX: openpyxl ile ilk worksheet; sütun başlıklarını otomatik bulur (ad, başlangıç, bitiş, vardiya).
  * PDF:  pdfplumber ile sayfa sayfa tablo + metin taraması.
  * Tarih formatları: DD.MM.YYYY, DD/MM/YYYY, YYYY-MM-DD. Tek bir gün verilmişse end_date = start_date.
  * MSSP dosyalarında A/B/C kolonu varsa shift_label olarak alınır.

Tüm endpoint'ler supervisor veya admin rolü ister (upload) / operatör+ (okuma).
"""
from __future__ import annotations
import io
import logging
import re
import uuid
from datetime import date, datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session

from ..auth import require_operator, require_supervisor
from ..database import get_db
from ..models import OnCallRoster, RosterTeam, User
from ..schemas import RosterEntryCreate, RosterEntryOut, RosterEntryUpdate, RosterUploadResult
from ..services import audit

log = logging.getLogger(__name__)
router = APIRouter(prefix="/roster", tags=["roster"])


# ---------- Date parsing helpers ----------

_DATE_PATTERNS = [
    "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y",
    "%d.%m.%y", "%d/%m/%y",
]


def _parse_date(raw) -> Optional[date]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    for fmt in _DATE_PATTERNS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # DD.MM (yıl eksik) — cari yılı kullan.
    m = re.match(r"^(\d{1,2})[./-](\d{1,2})$", s)
    if m:
        try:
            return date(datetime.now().year, int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None


def _norm_header(s: str) -> str:
    return (s or "").strip().lower().replace("i̇", "i")


_NAME_KEYS = {"ad", "ad soyad", "isim", "adi", "adı", "kisi", "kişi", "name", "person"}
_START_KEYS = {"baslangic", "başlangıç", "başlangic", "basla", "basladi", "start", "start_date", "from", "tarih başlangıç", "ilk"}
_END_KEYS = {"bitis", "bitiş", "bit", "end", "end_date", "to", "tarih bitiş", "son"}
_SHIFT_KEYS = {"vardiya", "shift", "shift_label", "etiket"}
_SINGLE_DATE_KEYS = {"tarih", "gun", "gün", "date"}


def _match(keys: set[str], header: str) -> bool:
    h = _norm_header(header)
    return h in keys


def _parse_xlsx(content: bytes, team: RosterTeam) -> tuple[list[dict], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # noqa: BLE001
        raise HTTPException(500, detail=f"openpyxl yüklü değil: {exc}")

    warnings: list[str] = []
    rows: list[dict] = []

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return rows, ["Çalışma sayfası bulunamadı."]

    header_row: Optional[list[str]] = None
    header_idx = 0
    # Başlık satırını (en fazla ilk 10 satır içinde) bul.
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i >= 10:
            break
        if r and any(cell for cell in r):
            header_row = [str(c) if c is not None else "" for c in r]
            header_idx = i
            break
    if header_row is None:
        return rows, ["Boş dosya."]

    name_i = start_i = end_i = shift_i = single_i = -1
    for idx, h in enumerate(header_row):
        if name_i < 0 and _match(_NAME_KEYS, h):
            name_i = idx
        elif start_i < 0 and _match(_START_KEYS, h):
            start_i = idx
        elif end_i < 0 and _match(_END_KEYS, h):
            end_i = idx
        elif shift_i < 0 and _match(_SHIFT_KEYS, h):
            shift_i = idx
        elif single_i < 0 and _match(_SINGLE_DATE_KEYS, h):
            single_i = idx

    if name_i < 0:
        warnings.append(f"Ad/isim sütunu tespit edilemedi — başlık: {header_row}")
        return rows, warnings

    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        if not r or not any(cell for cell in r):
            continue
        name = str(r[name_i]).strip() if name_i < len(r) and r[name_i] else ""
        if not name:
            continue
        start = _parse_date(r[start_i]) if 0 <= start_i < len(r) else None
        end = _parse_date(r[end_i]) if 0 <= end_i < len(r) else None
        if not start and 0 <= single_i < len(r):
            start = _parse_date(r[single_i])
            end = start
        if not end:
            end = start
        shift_label = None
        if 0 <= shift_i < len(r) and r[shift_i]:
            shift_label = str(r[shift_i]).strip()[:16]

        if not start:
            warnings.append(f"Satır {i + 1}: tarih okunamadı, atlandı.")
            continue

        rows.append({
            "team": team,
            "person_name": name[:255],
            "start_date": start,
            "end_date": end,
            "shift_label": shift_label,
            "notes": None,
        })
    return rows, warnings


# Basit PDF parser: satır bazlı metin taraması.
# Örn: "Ali Veli  01.05.2026 - 03.05.2026  A"
_PDF_LINE_RE = re.compile(
    r"^\s*(?P<name>[\wÇĞİÖŞÜçğıöşü .'-]+?)\s{2,}"
    r"(?P<start>\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?)"
    r"\s*[-–—]\s*(?P<end>\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?)"
    r"(?:\s+(?P<shift>[ABCabc]))?\s*$"
)


def _parse_pdf(content: bytes, team: RosterTeam) -> tuple[list[dict], list[str]]:
    try:
        import pdfplumber
    except ImportError as exc:  # noqa: BLE001
        raise HTTPException(500, detail=f"pdfplumber yüklü değil: {exc}")

    warnings: list[str] = []
    rows: list[dict] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for line in text.splitlines():
                m = _PDF_LINE_RE.match(line)
                if not m:
                    continue
                start = _parse_date(m.group("start"))
                end = _parse_date(m.group("end"))
                if not start or not end:
                    warnings.append(f"Sayfa {page_no}: tarih okunamadı — {line!r}")
                    continue
                shift = (m.group("shift") or "").upper() or None
                rows.append({
                    "team": team,
                    "person_name": m.group("name").strip()[:255],
                    "start_date": start,
                    "end_date": end,
                    "shift_label": shift,
                    "notes": None,
                })
    if not rows:
        warnings.append("PDF'ten satır çıkarılamadı. Beklenen format: 'Ad Soyad  GG.AA.YYYY - GG.AA.YYYY [A/B/C]'.")
    return rows, warnings


# ---------- Endpoints ----------

@router.get("", response_model=List[RosterEntryOut])
def list_roster(
    team: Optional[RosterTeam] = Query(None),
    start_from: Optional[date] = Query(None, description="Bu tarihten itibaren aktif olanlar"),
    start_to: Optional[date] = Query(None, description="Bu tarihe kadar aktif olanlar"),
    limit: int = Query(500, ge=1, le=2000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    _=Depends(require_operator),
):
    q = db.query(OnCallRoster)
    if team is not None:
        q = q.filter(OnCallRoster.team == team)
    if start_from:
        q = q.filter(OnCallRoster.end_date >= start_from)
    if start_to:
        q = q.filter(OnCallRoster.start_date <= start_to)
    return (
        q.order_by(OnCallRoster.start_date.asc(), OnCallRoster.id.asc())
        .offset(offset).limit(limit).all()
    )


@router.post("", response_model=RosterEntryOut, status_code=201)
def create_roster(
    payload: RosterEntryCreate,
    db: Session = Depends(get_db),
    current: User = Depends(require_supervisor),
):
    row = OnCallRoster(
        team=payload.team,
        person_name=payload.person_name,
        start_date=payload.start_date,
        end_date=payload.end_date,
        shift_label=payload.shift_label,
        notes=payload.notes,
        uploaded_by_id=current.id,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    audit(db, current, "roster.created", "roster", row.id, {"team": row.team.value})
    return row


@router.patch("/{roster_id}", response_model=RosterEntryOut)
def update_roster(
    roster_id: int,
    payload: RosterEntryUpdate,
    db: Session = Depends(get_db),
    current: User = Depends(require_supervisor),
):
    """Tek bir nöbet satırını günceller (ad, tarih aralığı, vardiya etiketi, not, ekip)."""
    row = db.query(OnCallRoster).filter(OnCallRoster.id == roster_id).first()
    if not row:
        raise HTTPException(404, detail="Kayıt bulunamadı")
    data = payload.model_dump(exclude_unset=True)

    # Tarih sırası tutarlılık kontrolü: end_date < start_date olamaz.
    new_start = data.get("start_date", row.start_date)
    new_end = data.get("end_date", row.end_date)
    if new_start and new_end and new_end < new_start:
        raise HTTPException(400, detail="Bitiş tarihi başlangıçtan önce olamaz.")

    for k, v in data.items():
        setattr(row, k, v)
    db.commit()
    db.refresh(row)
    audit(db, current, "roster.updated", "roster", row.id,
          {k: (v.isoformat() if hasattr(v, "isoformat") else str(v)) for k, v in data.items()})
    return row


@router.delete("/{roster_id}", status_code=204)
def delete_roster(
    roster_id: int,
    db: Session = Depends(get_db),
    current: User = Depends(require_supervisor),
):
    row = db.query(OnCallRoster).filter(OnCallRoster.id == roster_id).first()
    if not row:
        raise HTTPException(404, detail="Kayıt bulunamadı")
    db.delete(row)
    db.commit()
    audit(db, current, "roster.deleted", "roster", roster_id)
    return None


@router.delete("/batch/{upload_batch}", status_code=204)
def delete_roster_batch(
    upload_batch: str,
    db: Session = Depends(get_db),
    current: User = Depends(require_supervisor),
):
    rows = db.query(OnCallRoster).filter(OnCallRoster.upload_batch == upload_batch).all()
    if not rows:
        raise HTTPException(404, detail="Yükleme grubu bulunamadı")
    for r in rows:
        db.delete(r)
    db.commit()
    audit(db, current, "roster.batch_deleted", "roster", None,
          {"upload_batch": upload_batch, "count": len(rows)})
    return None


@router.post("/upload", response_model=RosterUploadResult)
async def upload_roster(
    team: RosterTeam = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: User = Depends(require_supervisor),
):
    """XLSX veya PDF dosyasını parse edip OnCallRoster satırları oluşturur.

    Kullanıcı tercihi: orijinal dosya saklanmaz — sadece normalize satırlar kaydedilir.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith((".xlsx", ".xlsm")):
        parsed, warnings = _parse_xlsx(content, team)
    elif filename.endswith(".pdf"):
        parsed, warnings = _parse_pdf(content, team)
    else:
        raise HTTPException(400, detail="Yalnızca .xlsx veya .pdf dosyaları kabul edilir.")

    if not parsed:
        return RosterUploadResult(
            upload_batch="",
            parsed_count=0,
            team=team,
            warnings=warnings or ["Hiç satır çıkarılamadı."],
        )

    batch_id = uuid.uuid4().hex
    for row in parsed:
        db.add(OnCallRoster(
            team=row["team"],
            person_name=row["person_name"],
            start_date=row["start_date"],
            end_date=row["end_date"],
            shift_label=row["shift_label"],
            notes=row["notes"],
            upload_batch=batch_id,
            uploaded_by_id=current.id,
        ))
    db.commit()
    audit(db, current, "roster.uploaded", "roster", None,
          {"team": team.value, "count": len(parsed), "batch": batch_id, "filename": file.filename})
    return RosterUploadResult(
        upload_batch=batch_id,
        parsed_count=len(parsed),
        team=team,
        warnings=warnings,
    )
